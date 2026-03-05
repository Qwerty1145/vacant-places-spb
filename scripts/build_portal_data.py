#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

URL_RE = re.compile(r"https?://[^\s<>\]\)\};,]+", re.IGNORECASE)

LABEL_DEADLINES = "Сроки переводов"
LABEL_DORM = "Общежитие"
LABEL_DORM_COMMENT = "Комментарий по общежитию"
LABEL_PROC = "Процедура перевода (кратко)"
LABEL_SPECIAL = "Особые условия перевода"
LABEL_DOCS = "Необходимые документы"
LABEL_VACANT = "Вакантные места для приема (перевода)"

PROFILE_OVERRIDE_FIELDS = {
    "src",
    "type",
    "summer",
    "winter",
    "notes",
    "dormInfo",
    "special",
    "proc",
    "docs",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def first_url(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    match = URL_RE.search(text)
    return match.group(0) if match else ""


def parse_dormitory_flag(value: str) -> bool | None:
    lowered = clean_text(value).lower()
    if not lowered:
        return None
    if lowered in {"+", "++", "есть", "да", "yes"}:
        return True
    if lowered in {"-", "нет", "no"}:
        return False
    return None


def append_text(base: str, addition: str) -> str:
    left = clean_text(base)
    right = clean_text(addition)
    if not right:
        return left
    if not left:
        return right
    if right in left:
        return left
    return f"{left} {right}"


def parse_profiles_workbook(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    result: dict[str, dict[str, str]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        item = {
            "name": sheet_name,
            "deadlines": "",
            "dorm_mark": "",
            "dorm_comment": "",
            "procedure": "",
            "special": "",
            "documents": "",
        }
        section: str | None = None

        for row in range(1, ws.max_row + 1):
            left = clean_text(ws.cell(row, 1).value)
            right = clean_text(ws.cell(row, 2).value)
            if not left and not right:
                continue

            if left.startswith("Профиль вуза:"):
                maybe_name = clean_text(left.replace("Профиль вуза:", "", 1))
                if maybe_name:
                    item["name"] = maybe_name
                section = None
                continue

            if left == LABEL_DEADLINES:
                item["deadlines"] = right
                section = None
                continue
            if left == LABEL_DORM:
                item["dorm_mark"] = right
                section = None
                continue
            if left == LABEL_DORM_COMMENT:
                item["dorm_comment"] = right
                section = None
                continue
            if left == LABEL_PROC:
                section = "procedure"
                continue
            if left == LABEL_SPECIAL:
                section = "special"
                continue
            if left == LABEL_DOCS:
                section = "documents"
                continue
            if left == LABEL_VACANT:
                section = None
                continue

            if section is None:
                continue

            candidate = left or right
            if not candidate:
                continue

            if (
                section == "special"
                and "особых условий не выделено" in candidate.lower()
            ):
                item["special"] = ""
                continue

            item[section] = append_text(item[section], candidate)

        result[item["name"]] = item

    return result


def parse_dates_workbook(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    headers = {
        col: clean_text(ws.cell(1, col).value).lower()
        for col in range(1, ws.max_column + 1)
    }

    def find_col(*terms: str) -> int | None:
        for col, header in headers.items():
            if all(term.lower() in header for term in terms):
                return col
        return None

    col_num = find_col("№") or find_col("n")
    col_name = find_col("краткое", "название", "вуз") or find_col("название", "вуз")
    col_dates = find_col("сроки", "подачи")
    col_link = find_col("ссылка", "вакант")
    col_note = find_col("примечания", "дате")
    col_dorm = find_col("общежитие")
    col_dorm_comment = find_col("комментарий", "общежити")

    if col_name is None:
        raise RuntimeError("Не найдена колонка с названием вуза в файле дат.")

    result: list[dict[str, Any]] = []
    seq = 0
    for row in range(2, ws.max_row + 1):
        name = clean_text(ws.cell(row, col_name).value)
        if not name:
            continue
        seq += 1

        source = ""
        if col_link is not None:
            cell = ws.cell(row, col_link)
            if cell.hyperlink and cell.hyperlink.target:
                source = clean_text(cell.hyperlink.target)
            if not source:
                source = first_url(cell.value)

        number = clean_text(ws.cell(row, col_num).value) if col_num is not None else str(seq)
        result.append(
            {
                "id": int(number) if number.isdigit() else seq,
                "name": name,
                "dates_raw": clean_text(ws.cell(row, col_dates).value) if col_dates else "",
                "dates_note": clean_text(ws.cell(row, col_note).value) if col_note else "",
                "source": source,
                "dorm_mark": clean_text(ws.cell(row, col_dorm).value) if col_dorm else "",
                "dorm_comment": clean_text(ws.cell(row, col_dorm_comment).value) if col_dorm_comment else "",
            }
        )

    return result


def load_profile_overrides(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    raw = payload.get("overrides", payload) if isinstance(payload, dict) else {}
    result: dict[str, dict[str, Any]] = {}
    if not isinstance(raw, dict):
        return result
    for key, value in raw.items():
        key_text = clean_text(key)
        if not key_text or not isinstance(value, dict):
            continue
        current: dict[str, Any] = {}
        for field, field_value in value.items():
            field_name = clean_text(field)
            if not field_name:
                continue
            if field_name not in PROFILE_OVERRIDE_FIELDS and field_name != "dorm":
                continue
            current[field_name] = field_value
        if current:
            result[key_text] = current
    return result


def parse_type_map_from_template(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8")
    start = text.find("const UNIS=[")
    if start < 0:
        return {}
    end = text.find("];", start)
    chunk = text[start:end] if end > start else text[start:]

    mapping: dict[str, str] = {}
    pattern = re.compile(
        r"name:`([^`]+)`\s*,\s*src:`[^`]*`\s*,\s*type:`([^`]+)`",
        re.S,
    )
    for match in pattern.finditer(chunk):
        name = clean_text(match.group(1))
        uni_type = clean_text(match.group(2))
        if name and uni_type:
            mapping[name] = uni_type
    return mapping


def parse_int(value: Any) -> int:
    text = clean_text(value)
    if not text:
        return 0
    numbers = re.findall(r"\d+", text)
    if not numbers:
        return 0
    ints = [int(number) for number in numbers]
    if len(ints) > 1 and len(set(ints)) == 1:
        return ints[0]
    return sum(ints)


def normalize_course(value: Any) -> Any:
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    if match:
        return int(match.group(0))
    return text


def rows_to_portal_format(rows: list[list[Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        current = [clean_text(cell) for cell in row]
        if len(current) < 8:
            current.extend([""] * (8 - len(current)))
        code, direction, program, level, course, form, budget, paid = current[:8]
        budget_num = parse_int(budget)
        paid_num = parse_int(paid)

        if not any([code, direction, program, level, course, form, budget_num, paid_num]):
            continue

        result.append(
            {
                "code": code,
                "dir": direction,
                "program": program,
                "level": level,
                "course": normalize_course(course),
                "form": form,
                "budget": budget_num,
                "paid": paid_num,
            }
        )
    return result


def load_vacancy_data(path: Path) -> dict[str, dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    universities = payload.get("universities", [])
    by_name: dict[str, dict[str, Any]] = {}
    for university in universities:
        name = clean_text(university.get("university"))
        vacancies = university.get("vacancies") or {}
        rows = vacancies.get("standardized_rows") or []
        by_name[name] = {
            "status": clean_text(vacancies.get("status")) or "error",
            "message": clean_text(vacancies.get("message") or vacancies.get("note")),
            "source": clean_text(vacancies.get("source_url") or university.get("official_vacancies_link")),
            "rows": rows_to_portal_format(rows),
            "dorm_flag": university.get("dormitory", {}).get("has_dormitory"),
            "dorm_details": clean_text(university.get("dormitory", {}).get("details")),
        }
    return by_name


def load_season_split_fn(build_site_path: Path):
    spec = importlib.util.spec_from_file_location("build_site_module", build_site_path)
    if spec is None or spec.loader is None:
        return None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return getattr(module, "season_split", None)


def split_deadlines(raw_text: str, season_split_fn: Any) -> tuple[str, str, str]:
    text = clean_text(raw_text)
    if not text:
        return "", "", ""

    if callable(season_split_fn):
        parts = season_split_fn(text)
        summer = "; ".join(parts.get("summer", []))
        winter = "; ".join(parts.get("winter", []))
        other = "; ".join(parts.get("other", []))
        return summer, winter, other

    chunks = [clean_text(chunk) for chunk in re.split(r"[;\n]+", text) if clean_text(chunk)]
    summer_list: list[str] = []
    winter_list: list[str] = []
    other_list: list[str] = []
    for chunk in chunks:
        lowered = chunk.lower()
        if any(token in lowered for token in ("янв", "фев", "дек", "зим", "весенн")):
            winter_list.append(chunk)
        elif any(token in lowered for token in ("июн", "июл", "авг", "сен", "окт", "ноя", "лет", "осенн")):
            summer_list.append(chunk)
        else:
            other_list.append(chunk)
    return "; ".join(summer_list), "; ".join(winter_list), "; ".join(other_list)


def build_payload(
    profiles_path: Path,
    dates_path: Path,
    vacancies_path: Path,
    template_path: Path,
    build_site_path: Path,
    profile_overrides_path: Path,
) -> dict[str, Any]:
    profiles = parse_profiles_workbook(profiles_path)
    dates = parse_dates_workbook(dates_path)
    vacancies = load_vacancy_data(vacancies_path)
    type_map = parse_type_map_from_template(template_path)
    season_split_fn = load_season_split_fn(build_site_path)
    profile_overrides = load_profile_overrides(profile_overrides_path)

    unis: list[dict[str, Any]] = []
    vacancies_by_id: dict[str, list[dict[str, Any]]] = {}
    total_rows = 0
    loaded_unis = 0

    for fallback_id, date_item in enumerate(dates, start=1):
        uni_id = int(date_item.get("id") or fallback_id)
        name = clean_text(date_item.get("name"))
        profile = profiles.get(name, {})
        vacancy = vacancies.get(name, {})

        deadlines_raw = (
            clean_text(profile.get("deadlines"))
            or clean_text(date_item.get("dates_raw"))
        )
        summer, winter, other = split_deadlines(deadlines_raw, season_split_fn)
        summer_text = summer or deadlines_raw or "Уточнять на официальном сайте"
        winter_text = winter or "Уточнять на официальном сайте"

        notes = clean_text(date_item.get("dates_note"))
        if other:
            notes = append_text(notes, other)

        dorm_flag = parse_dormitory_flag(profile.get("dorm_mark", ""))
        if dorm_flag is None:
            dorm_flag = parse_dormitory_flag(clean_text(date_item.get("dorm_mark")))
        if dorm_flag is None and isinstance(vacancy.get("dorm_flag"), bool):
            dorm_flag = vacancy.get("dorm_flag")

        dorm_info = (
            clean_text(profile.get("dorm_comment"))
            or clean_text(date_item.get("dorm_comment"))
            or clean_text(vacancy.get("dorm_details"))
        )
        src = (
            clean_text(date_item.get("source"))
            or clean_text(vacancy.get("source"))
        )

        rows = vacancy.get("rows", [])
        status = clean_text(vacancy.get("status")) or ("ok" if rows else "error")
        message = clean_text(vacancy.get("message"))
        if status == "ok" and not rows:
            status = "error"
            base_message = "Таблица извлечена без пригодных строк (требуется ручная проверка источника)."
            message = f"{base_message} {message}".strip()

        row_count = len(rows)
        total_rows += row_count
        if status == "ok" and row_count > 0:
            loaded_unis += 1

        uni = {
            "id": uni_id,
            "abbr": name,
            "name": name,
            "src": src,
            "type": type_map.get(name, "federal"),
            "summer": summer_text,
            "winter": winter_text,
            "notes": notes,
            "dorm": bool(dorm_flag) if dorm_flag is not None else False,
            "dormInfo": dorm_info,
            "special": clean_text(profile.get("special")),
            "proc": clean_text(profile.get("procedure")),
            "docs": clean_text(profile.get("documents")),
            "vacancyStatus": status,
            "vacancyMessage": message,
            "vacancySource": clean_text(vacancy.get("source")) or src,
            "vacancyRowCount": row_count,
        }

        override = profile_overrides.get(str(uni_id)) or profile_overrides.get(name)
        if isinstance(override, dict):
            for field, value in override.items():
                if field == "dorm":
                    parsed = parse_dormitory_flag(value)
                    if parsed is not None:
                        uni["dorm"] = bool(parsed)
                    continue
                cleaned = clean_text(value)
                if cleaned and field in PROFILE_OVERRIDE_FIELDS:
                    uni[field] = cleaned

        unis.append(uni)
        vacancies_by_id[str(uni_id)] = rows

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "stats": {
            "totalUniversities": len(unis),
            "loadedUniversities": loaded_unis,
            "errorUniversities": sum(
                1 for uni in unis if uni["vacancyStatus"] == "error"
            ),
            "totalVacancyRows": total_rows,
        },
        "unis": unis,
        "vacanciesById": vacancies_by_id,
    }
    return payload


def write_portal_data_js(payload: dict[str, Any], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    content = "window.PORTAL_DATA = " + json.dumps(payload, ensure_ascii=False) + ";\n"
    out_path.write_text(content, encoding="utf-8")


def build_portal_html_from_template(template_path: Path, out_html_path: Path) -> None:
    content = template_path.read_text(encoding="utf-8")
    marker = "<script>\nconst {useState,useMemo}=React;"
    start = content.find(marker)
    if start < 0:
        raise RuntimeError("Не найден inline-скрипт в шаблоне портала.")
    end = content.find("</script>", start)
    if end < 0:
        raise RuntimeError("Не найдено закрытие inline-скрипта в шаблоне портала.")

    replacement = (
        '<script defer src="./data/portal_data.js"></script>\n'
        '<script defer src="./portal_app.js"></script>'
    )
    rebuilt = content[:start] + replacement + content[end + len("</script>") :]

    out_html_path.parent.mkdir(parents=True, exist_ok=True)
    out_html_path.write_text(rebuilt, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build portal data and HTML for spb_transfer_portal."
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=Path("/mnt/c/Users/arsen/Downloads/spb_transfer_portal (1).html"),
    )
    parser.add_argument(
        "--profiles",
        type=Path,
        default=Path("/mnt/c/Users/arsen/Downloads/transfer_profiles_by_university.xlsx"),
    )
    parser.add_argument(
        "--dates",
        type=Path,
        default=Path("/mnt/c/Users/arsen/Downloads/transfer_dates_spb_FULL_v2.xlsx"),
    )
    parser.add_argument(
        "--vacancies",
        type=Path,
        default=Path("site/data/universities.json"),
    )
    parser.add_argument(
        "--build-site-script",
        type=Path,
        default=Path("scripts/build_site.py"),
    )
    parser.add_argument(
        "--out-html",
        type=Path,
        default=Path("site/spb_transfer_portal.html"),
    )
    parser.add_argument(
        "--out-data",
        type=Path,
        default=Path("site/data/portal_data.js"),
    )
    parser.add_argument(
        "--profile-overrides",
        type=Path,
        default=Path("site/data/manual_profile_overrides.json"),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    payload = build_payload(
        profiles_path=args.profiles,
        dates_path=args.dates,
        vacancies_path=args.vacancies,
        template_path=args.template,
        build_site_path=args.build_site_script,
        profile_overrides_path=args.profile_overrides,
    )
    write_portal_data_js(payload, args.out_data)
    build_portal_html_from_template(args.template, args.out_html)

    stats = payload.get("stats", {})
    print(
        "Saved:",
        args.out_html,
        "|",
        args.out_data,
        f"| universities={stats.get('totalUniversities', 0)}",
        f"| loaded={stats.get('loadedUniversities', 0)}",
        f"| errors={stats.get('errorUniversities', 0)}",
        f"| vacancy_rows={stats.get('totalVacancyRows', 0)}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
