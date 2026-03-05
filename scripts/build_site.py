#!/usr/bin/env python3
from __future__ import annotations

import argparse
import concurrent.futures as futures
import io
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    openpyxl = None
try:
    import pandas as pd
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    pd = None
try:
    import pdfplumber
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    pdfplumber = None
try:
    import requests
    from requests import Response
    from requests.exceptions import RequestException, SSLError
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    requests = None
    Response = Any  # type: ignore
    RequestException = Exception  # type: ignore
    SSLError = Exception  # type: ignore
try:
    import urllib3
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    urllib3 = None
try:
    from bs4 import BeautifulSoup
    from bs4 import UnicodeDammit
    BS4_AVAILABLE = True
except ImportError:  # pragma: no cover - optional for non-scrape workflows
    BeautifulSoup = Any  # type: ignore
    UnicodeDammit = None
    BS4_AVAILABLE = False

if urllib3 is not None:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

EXCLUDED_MARKERS = {"-"}
MAX_ROWS_PER_TABLE = 1200
REQUEST_TIMEOUT = (10, 25)
TABLE_KEYWORDS = (
    "вакант",
    "мест",
    "бюджет",
    "направлен",
    "специальн",
    "курс",
    "очная",
    "заочная",
    "колич",
    "финанс",
    "обуч",
)
DOWNLOAD_KEYWORDS = ("вакант", "vacant", "vakant", "budget", "перевод")
FILE_EXTENSIONS = {".pdf", ".xls", ".xlsx", ".csv"}
URL_RE = re.compile(r"https?://[^\s<>\]\)\};,]+", re.IGNORECASE)
CYRILLIC_RE = re.compile(r"[А-Яа-яЁё]")
CODE_RE = re.compile(r"\b\d{1,2}\.\d{1,2}\.\d{1,2}\b")
HEADER_HINTS = ("код", "наименован", "направлен", "специальн", "форма", "курс", "уровень", "бюджет")
SIGNATURE_HINTS = ("документ подписан", "уникальный программный ключ", "должность")

MENU_WORDS = (
    "факультет",
    "институт",
    "библиот",
    "журнал",
    "расписан",
    "подраздел",
    "отдел",
    "служб",
    "центр",
    "деятельность",
    "консультац",
)

HIDDEN_CELL_SELECTORS = ".hide, .element-invisible, .sr-only, .visually-hidden"
HIDDEN_CELL_CLASSES = {"hide", "sr-only", "visually-hidden"}

STANDARD_COLUMNS = [
    "Код специальности",
    "Направление подготовки",
    "Программа",
    "Уровень образования",
    "Курс",
    "Форма обучения",
    "Бюджетные места",
    "Платные места",
]

# Use only targeted, same-university overrides for unstable mirror aliases.
SOURCE_OVERRIDES: dict[str, str] = {
    "u027": "https://rshu.ru/sveden/vacant/",
    "u017": "https://www.gikit.ru/sveden/vacant/",
}

MANUAL_VACANCY_FALLBACKS: dict[str, dict[str, Any]] = {
    "u010": {
        "source_url": "https://doc.spbgasu.ru/vacant/BFVacant.pdf",
        "message": (
            "Таблица заполнена вручную по официальному PDF (данные на 16.01.2026)."
        ),
        "columns": [
            "Факультет",
            "Специальность",
            "1 курс",
            "2 курс",
            "3 курс",
            "4 курс",
            "5 курс",
            "6 курс",
        ],
        "rows": [
            ["Автомобильно-дорожный", "23.05.01 Наземные транспортно-технологические средства", "5", "5", "8", "9", "27", ""],
            ["Автомобильно-дорожный", "15.03.03 Прикладная механика", "0", "12", "4", "34", "", ""],
            ["Автомобильно-дорожный", "23.03.01 Технология транспортных процессов", "2", "14", "15", "13", "", ""],
            ["Автомобильно-дорожный", "23.03.03 Эксплуатация транспортно-технологических машин и комплексов", "5", "19", "17", "24", "", ""],
            ["Автомобильно-дорожный", "08.03.01 Строительство", "3", "7", "20", "4", "", ""],
            ["Автомобильно-дорожный", "08.04.01 Строительство", "3", "14", "", "", "", ""],
            ["Автомобильно-дорожный", "08.05.01 Строительство уникальных зданий и сооружений", "0", "", "", "0", "3", "0"],
            ["Автомобильно-дорожный", "15.03.06 Механика и робототехника", "0", "0", "", "", "", ""],
            ["Автомобильно-дорожный", "23.04.01 Технология транспортных процессов (магистратура)", "0", "0", "", "", "", ""],
            ["Автомобильно-дорожный", "15.04.03 Прикладная механика (магистратура)", "0", "2", "", "", "", ""],
            ["Автомобильно-дорожный", "23.04.02 Наземные транспортно-технологические комплексы (магистратура)", "0", "1", "", "", "", ""],
            ["Автомобильно-дорожный", "23.04.03 Эксплуатация транспортно-технологических машин и комплексов (магистратура)", "0", "0", "", "", "", ""],
            ["Архитектурный", "07.03.01 Архитектура", "0", "1", "0", "1", "1", ""],
            ["Архитектурный", "07.03.04 Градостроительство", "0", "1", "1", "1", "0", ""],
            ["Архитектурный", "07.03.03 Дизайн архитектурной среды", "0", "1", "1", "1", "0", ""],
            ["Архитектурный", "07.03.02 Реконструкция и реставрация архитектурного наследия", "0", "2", "2", "2", "5", ""],
            ["Архитектурный", "35.03.10 Ландшафтная архитектура", "0", "4", "7", "7", "", ""],
            ["Архитектурный", "54.03.01 Дизайн", "0", "0", "0", "", "", ""],
            ["Архитектурный", "54.05.03 Графика", "0", "0", "0", "", "", ""],
            ["Архитектурный", "07.04.03 Дизайн архитектурной среды (магистратура)", "4", "7", "", "", "", ""],
            ["Архитектурный", "07.04.02 Реконструкция и реставрация архитектурного наследия (магистратура)", "1", "5", "", "", "", ""],
            ["Архитектурный", "07.04.01 Архитектура (магистратура)", "7", "12", "", "", "", ""],
            ["Архитектурный", "35.04.09 Ландшафтная архитектура (магистратура)", "", "0", "", "", "", ""],
            ["Архитектурный", "07.04.04 Градостроительство (магистратура)", "2", "15", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "21.03.02 Землеустройство и кадастры", "4", "4", "4", "7", "", ""],
            ["Инженерной экологии и городского хозяйства", "21.04.02 Землеустройство и кадастры (магистратура)", "1", "1", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "13.03.01 Теплоэнергетика и теплотехника", "", "6", "", "2", "", ""],
            ["Инженерной экологии и городского хозяйства", "13.04.01 Теплоэнергетика и теплотехника (магистратура)", "2", "9", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "13.03.02 Электроэнергетика и электротехника", "0", "", "", "2", "", ""],
            ["Инженерной экологии и городского хозяйства", "13.04.02 Электроэнергетика и электротехника", "1", "", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "08.03.01 Строительство", "4", "9", "15", "26", "", ""],
            ["Инженерной экологии и городского хозяйства", "08.04.01 Строительство (магистратура)", "1", "15", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "09.03.02 Информационные системы и технологии", "1", "2", "2", "0", "", ""],
            ["Инженерной экологии и городского хозяйства", "09.04.02 Информационные системы и технологии", "0", "7", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "20.03.02 Природообустройство и водопользование", "1", "", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "20.04.02 Природообустройство и водопользование", "0", "", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "21.05.03 Прикладная геодезия", "0", "", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "09.03.03 Прикладная информатика", "0", "0", "0", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "09.04.03 Прикладная информатика", "1", "0", "", "", "", ""],
            ["Инженерной экологии и городского хозяйства", "01.03.02 Прикладная математика и информатика", "2", "5", "1", "5", "", ""],
            ["Инженерной экологии и городского хозяйства", "01.04.02 Прикладная математика и информатика (магистратура)", "1", "0", "", "", "", ""],
            ["Строительный", "27.03.01 Стандартизация и метрология", "", "", "", "3", "", ""],
            ["Строительный", "27.04.01 Стандартизация и метрология (магистратура)", "", "8", "", "", "", ""],
            ["Строительный", "08.03.01 Строительство", "5", "15", "34", "32", "", ""],
            ["Строительный", "08.04.01 Строительство (магистратура)", "2", "25", "", "", "", ""],
            ["Строительный", "08.05.01 Строительство уникальных зданий и сооружений", "0", "1", "0", "1", "1", "0"],
            ["Строительный", "20.03.01 Техносферная безопасность", "0", "9", "7", "9", "", ""],
            ["Факультет экономики и управления", "38.03.02 Менеджмент", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.1 Строительные конструкции, здания и сооружения", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.12 Архитектура зданий и сооружений. Творческие концепции архитектурной деятельности", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.11 Теория и история архитектуры, реставрация и реконструкция историко-архитектурного наследия", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.13 Градостроительство, планировка сельских населенных пунктов", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.3 Теплоснабжение, вентиляция, кондиционирование воздуха, газоснабжение и освещение", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.2 Основания и фундаменты, подземные сооружения", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.8 Проектирование и строительство дорог, метрополитенов, аэродромов, мостов и транспортных тоннелей", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.9 Строительная механика", "0", "1", "0", "0", "", ""],
            ["Аспирантура", "2.1.4 Водоснабжение, канализация, строительные системы охраны водных ресурсов", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.5 Строительные материалы и изделия", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.1.7 Технология и организация строительства", "0", "0", "0", "0", "", ""],
            ["Аспирантура", "2.5.11 Наземные транспортно-технологические средства и комплексы", "0", "0", "1", "1", "", ""],
            ["Аспирантура", "2.9.5 Эксплуатация автомобильного транспорта", "1", "3", "0", "0", "", ""],
            ["Аспирантура", "1.2.2 Математическое моделирование, численные методы и комплексы программ", "1", "0", "0", "0", "", ""],
            ["Аспирантура", "5.6.6 История науки и техники", "", "", "", "", "", ""],
            ["Аспирантура", "5.2.6 Менеджмент", "1", "1", "", "", "", ""],
            ["Аспирантура", "5.2.3 Региональная и отраслевая экономика", "0", "1", "0", "", "", ""],
        ],
    },
    "u040": {
        "source_url": "https://igps.ru/sveden/vacant/?special=Y",
        "message": (
            "Данные таблицы добавлены из доступной копии официальной страницы "
            "в режиме 'special=Y' (из-за сетевой недоступности домена при прямом запросе)."
        ),
        "columns": [
            "Код",
            "Направление подготовки",
            "Уровень",
            "Профиль",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            [
                "20.05.01",
                "Пожарная безопасность",
                "Специалитет",
                "Пожарная безопасность",
                "2",
                "очная",
                "0",
                "0",
                "0",
                "13",
            ]
        ],
    },
    "u044": {
        "source_url": "https://spbu.fsin.gov.ru/sveden/vacant/",
        "message": (
            "Таблица собрана из официальных PDF на странице вуза "
            "(доступ к домену ограничен из текущей сети)."
        ),
        "columns": [
            "Документ",
            "Код",
            "Направление подготовки",
            "Уровень",
            "Профиль",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            [
                "vakmest20.pdf",
                "40.05.02",
                "Правоохранительная деятельность",
                "Специалитет",
                "Оперативно-розыскная деятельность",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ],
            [
                "vakmest20_1.pdf",
                "40.05.02",
                "Правоохранительная деятельность",
                "Специалитет",
                "Оперативно-розыскная деятельность",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ],
        ],
    },
    "u045": {
        "source_url": "https://university.mvd.ru/sveden/vacant?special=Y",
        "message": (
            "На официальной странице в режиме 'special=Y' доступны заголовки раздела, "
            "но строки таблицы вакантных мест не опубликованы в статической разметке."
        ),
        "columns": ["Статус публикации"],
        "rows": [
            [
                "Строки таблицы вакантных мест не представлены в открытой статической версии страницы.",
            ]
        ],
    },
    "u050": {
        "source_url": (
            "https://vma.mil.ru/Obrazovanie/"
            "Vakantnie-mesta-dlya-priema-perevoda-obuchayushhihsya"
        ),
        "message": (
            "Данные взяты из официальной страницы в доступной веб-копии "
            "(прямой доступ к домену из текущей сети недоступен)."
        ),
        "columns": [
            "Код",
            "Направление подготовки",
            "Уровень",
            "Профиль",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            [
                "26.05.04",
                "Применение и эксплуатация технических систем надводных кораблей и подводных лодок",
                "Специалитет",
                "Применение и эксплуатация технических систем надводных кораблей и подводных лодок",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ]
        ],
    },
    "u051": {
        "source_url": "https://vamto.mil.ru/Obrazovanie/Vakantnye-mesta-dlya-priema-perevoda",
        "message": (
            "Данные взяты из официальной страницы в доступной веб-копии "
            "(прямой доступ к домену из текущей сети недоступен)."
        ),
        "columns": [
            "Код",
            "Направление подготовки",
            "Уровень",
            "Профиль",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            [
                "56.05.01",
                "Тыловое обеспечение",
                "Специалитет",
                "Организация обеспечения войск (сил)",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ],
            [
                "56.05.02",
                "Военно-политическая работа",
                "Специалитет",
                "Военно-политическая работа в войсках (силах)",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ],
            [
                "40.05.01",
                "Правовое обеспечение национальной безопасности",
                "Специалитет",
                "Уголовно-правовая специализация",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "0",
            ],
        ],
    },
    "u052": {
        "source_url": (
            "https://mvaa.mil.ru/Obrazovanie/"
            "Vakantnie-mesta-dlya-priema-perevoda-obuchayushhihsya"
        ),
        "message": (
            "На официальной странице указано: «Информация отсутствует» "
            "(данные получены из доступной веб-копии)."
        ),
        "columns": ["Статус публикации"],
        "rows": [["Информация отсутствует"]],
    },
    "u053": {
        "source_url": "https://spvi.rosguard.gov.ru/sveden/vacant?special=Y",
        "message": (
            "Данные взяты из официальной страницы в режиме 'special=Y' "
            "(прямой доступ к домену из текущей сети недоступен)."
        ),
        "columns": [
            "Тип таблицы",
            "Код",
            "Направление подготовки",
            "Уровень",
            "Профиль",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            [
                "Прием",
                "38.05.02",
                "Таможенное дело",
                "Специалитет",
                "Таможенные платежи и валютное регулирование",
                "-",
                "очная",
                "0",
                "0",
                "0",
                "15",
            ],
            [
                "Перевод",
                "38.05.02",
                "Таможенное дело",
                "Специалитет",
                "Таможенные платежи и валютное регулирование",
                "1",
                "очная",
                "0",
                "0",
                "0",
                "15",
            ],
        ],
    },
    "u031": {
        "source_url": "https://spbguga.ru/ru/sveden/vacant/",
        "message": "Данные вручную извлечены с официальной страницы (таблица вакантных мест).",
        "columns": [
            "Код",
            "Наименование",
            "Профиль",
            "Уровень образования",
            "Курс",
            "Форма",
            "Фед. бюджет",
            "Бюджет субъекта РФ",
            "Местный бюджет",
            "Платное обучение",
        ],
        "rows": [
            ["38.03.01", "Экономика", "-", "Высшее образование — бакалавриат", "1", "Очная", "1", "0", "0", "109"],
            ["38.03.02", "Менеджмент", "-", "Высшее образование — бакалавриат", "1", "Очная", "2", "0", "0", "80"],
            ["38.03.02", "Менеджмент", "-", "Высшее образование — бакалавриат", "2", "Очная", "1", "0", "0", "79"],
            ["38.03.02", "Менеджмент", "-", "Высшее образование — бакалавриат", "3", "Очная", "0", "0", "0", "79"],
            ["38.03.02", "Менеджмент", "-", "Высшее образование — бакалавриат", "4", "Очная", "0", "0", "0", "80"],
            ["38.03.02", "Менеджмент", "-", "Высшее образование — бакалавриат", "5", "Очная", "0", "0", "0", "80"],
            ["20.03.01", "Техносферная безопасность", "-", "Высшее образование — бакалавриат", "1", "Очная", "13", "0", "0", "0"],
            ["38.04.01", "Экономика", "-", "Высшее образование — магистратура", "1", "Очная", "0", "0", "0", "25"],
            ["38.04.02", "Менеджмент", "-", "Высшее образование — магистратура", "1", "Очная", "0", "0", "0", "49"],
            ["38.04.02", "Менеджмент", "-", "Высшее образование — магистратура", "2", "Очная", "0", "0", "0", "49"],
            ["38.04.08", "Финансы и кредит", "-", "Высшее образование — магистратура", "1", "Очная", "0", "0", "0", "21"],
            ["38.04.08", "Финансы и кредит", "-", "Высшее образование — магистратура", "2", "Очная", "0", "0", "0", "22"],
            ["20.04.01", "Техносферная безопасность", "-", "Высшее образование — магистратура", "1", "Очная", "5", "0", "0", "0"],
            ["20.04.01", "Техносферная безопасность", "-", "Высшее образование — магистратура", "2", "Очная", "4", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "1", "Очная", "10", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "2", "Очная", "10", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "3", "Очная", "10", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "4", "Очная", "10", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "5", "Очная", "20", "0", "0", "0"],
            ["25.05.05", "Эксплуатация воздушных судов и организация воздушного движения", "-", "Высшее образование — специалитет", "6", "Очная", "15", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "1", "Очная", "20", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "2", "Очная", "20", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "3", "Очная", "20", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "4", "Очная", "20", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "5", "Очная", "30", "0", "0", "0"],
            ["25.05.04", "Летная эксплуатация и применение авиационных комплексов", "-", "Высшее образование — специалитет", "6", "Очная", "20", "0", "0", "0"],
        ],
    },
}

SESSION = requests.Session() if requests is not None else None
if SESSION is not None:
    SESSION.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/130.0.0.0 Safari/537.36"
            )
        }
    )


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = recover_mojibake(str(value))
    text = text.replace("в„–", "№")
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def text_quality_score(text: str) -> int:
    cyrillic_count = len(CYRILLIC_RE.findall(text))
    latin_mojibake_markers = sum(text.count(ch) for ch in ("Ð", "Ñ", "Ã", "Ê", "Â", "�"))
    utf8_cp1251_markers = len(re.findall(r"[РС][А-Яа-яЁё]", text))
    box_drawing = sum(1 for ch in text if 0x2500 <= ord(ch) <= 0x259F)
    math_ops = sum(1 for ch in text if 0x2200 <= ord(ch) <= 0x22FF)
    letters = [ch for ch in text if ch.isalpha()]
    upper = sum(1 for ch in letters if ch.isupper())
    lower = sum(1 for ch in letters if ch.islower())
    upper_ratio = upper / (upper + lower) if (upper + lower) else 0.0
    upper_penalty = int(upper_ratio * 8) if upper and lower and upper_ratio > 0.4 else 0

    return (
        cyrillic_count
        - (latin_mojibake_markers * 2)
        - utf8_cp1251_markers
        - (box_drawing * 3)
        - (math_ops * 2)
        - (text.count("�") * 4)
        - upper_penalty
    )


def normalize_header(text: str) -> str:
    return re.sub(r"\s+", " ", clean_text(text).lower())


def parse_number(value: str) -> int:
    text = clean_text(value)
    if not text or text in {"-", "—", "нет", "нет данных"}:
        return 0
    numbers = re.findall(r"\d+", text)
    if not numbers:
        return 0
    ints = [int(num) for num in numbers]
    # Some sources duplicate the same number inside a single cell (e.g. "2 2"). Treat as one value.
    if len(ints) > 1 and len(set(ints)) == 1:
        return ints[0]
    return sum(ints)


def normalize_course(value: str) -> str:
    text = clean_text(value)
    if text in {"-", "—", "–"}:
        return ""
    lowered = text.lower()
    if "все" in lowered and "курс" in lowered:
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else text


def parse_int_token(value: str) -> int | None:
    token = clean_text(value)
    if not token:
        return None
    token = token.replace(",", ".")
    if not re.fullmatch(r"\d+(?:\.\d+)?", token):
        return None
    try:
        return int(float(token))
    except ValueError:
        return None


def parse_course_tokens(value: str) -> list[str]:
    text = clean_text(value)
    if not text or text in {"-", "—", "–"}:
        return []

    direct = parse_int_token(text)
    if direct is not None:
        return [str(direct)]

    parts = [part for part in re.split(r"\s+", text) if part]
    if parts and all(re.fullmatch(r"[-—–]|\d+(?:[.,]\d+)?", part) for part in parts):
        tokens: list[str] = []
        for part in parts:
            if re.fullmatch(r"[-—–]", part):
                continue
            parsed = parse_int_token(part)
            if parsed is not None:
                tokens.append(str(parsed))
        return tokens

    raw_numbers = re.findall(r"\d+(?:[.,]\d+)?", text)
    parsed_numbers = []
    for item in raw_numbers:
        parsed = parse_int_token(item)
        if parsed is not None:
            parsed_numbers.append(str(parsed))
    return parsed_numbers


def parse_amount_tokens(value: str) -> list[int]:
    text = clean_text(value)
    if not text or text.lower() in {"нет", "нет данных"}:
        return []

    direct = parse_int_token(text)
    if direct is not None:
        return [direct]

    parts = [part for part in re.split(r"\s+", text) if part]
    if parts and all(re.fullmatch(r"[-—–]|\d+(?:[.,]\d+)?", part) for part in parts):
        values: list[int] = []
        for part in parts:
            if re.fullmatch(r"[-—–]", part):
                values.append(0)
                continue
            parsed = parse_int_token(part)
            if parsed is not None:
                values.append(parsed)
        return values

    raw_tokens = re.findall(r"\d+(?:[.,]\d+)?|[-—–]", text)
    if not raw_tokens:
        number = parse_number(text)
        if number == 0 and not re.search(r"\d", text):
            return []
        return [number]
    values: list[int] = []
    for token in raw_tokens:
        parsed = parse_int_token(token)
        if parsed is not None:
            values.append(parsed)
        else:
            values.append(0)
    return values


def detect_form_literal(text: str) -> str:
    lowered = normalize_header(text)
    has_full = "очн" in lowered
    has_part = "заоч" in lowered
    if has_full and has_part and "очно-заоч" not in lowered and "очно заоч" not in lowered:
        return ""
    if "очно-заоч" in lowered or "очно заоч" in lowered:
        return "очно-заочная"
    if "заоч" in lowered:
        return "заочная"
    if "очн" in lowered:
        return "очная"
    if "дистанц" in lowered:
        return "дистанционная"
    if "вечер" in lowered:
        return "вечерняя"
    return ""


def split_code_and_direction(value: str) -> tuple[str, str] | None:
    text = clean_text(value)
    if not text:
        return None
    # Only split when code and direction are separated by whitespace or dash.
    # Otherwise strings like "01.03.02" would backtrack and split into ("01.03.0", "2").
    match = re.match(
        r"^(\d{1,2}\.\d{1,2}\.\d{1,2})(?:\s*[–—-]\s*|\s+)(.+)$",
        text,
    )
    if not match:
        return None
    code = match.group(1)
    direction = clean_text(match.group(2))
    if not direction:
        return None
    if not re.search(r"[A-Za-zА-Яа-яЁё]", direction):
        return None
    return code, direction


def looks_like_level_value(value: str) -> bool:
    lowered = value.lower()
    return any(
        term in lowered
        for term in (
            "бакалав",
            "магистр",
            "специал",
            "аспиран",
            "ординат",
            "среднее профессион",
            "спо",
            "базовое высшее",
        )
    )


def infer_level_from_code(code: str) -> str:
    text = clean_text(code)
    match = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{1,2})", text)
    if not match:
        return ""

    first = int(match.group(1))
    middle = int(match.group(2))

    if middle == 2:
        return "Среднее профессиональное образование"
    if middle == 3:
        return "Высшее образование - бакалавриат"
    if middle == 4:
        return "Высшее образование - магистратура"
    if middle == 5:
        return "Высшее образование - специалитет"
    if middle == 6:
        return "Высшее образование - подготовка кадров высшей квалификации"
    if middle in {7, 8, 9} and first < 10:
        # Научные специальности формата 1.6.16, 5.2.3 и т.п.
        return "Высшее образование - подготовка кадров высшей квалификации"
    return ""


def is_summary_value(value: str) -> bool:
    lowered = clean_text(value).lower()
    if not lowered:
        return False
    if lowered in {"-", "—", "–"}:
        return True
    if lowered.startswith("итог") or lowered.startswith("всего"):
        return True
    if "все" in lowered and "курс" in lowered:
        return True
    return False


def is_numeric_cell(value: str) -> bool:
    text = clean_text(value)
    return bool(text) and bool(re.fullmatch(r"\d+(?:\.\d+)?", text))


def is_valid_code(value: str) -> bool:
    text = clean_text(value)
    return bool(text) and bool(re.fullmatch(r"\d{1,2}\.\d{1,2}\.\d{1,2}", text))


def best_index(columns: list[str], scorer: Any) -> int | None:
    best = None
    best_score = 0
    for idx, column in enumerate(columns):
        score = scorer(normalize_header(column))
        if score > best_score:
            best_score = score
            best = idx
    return best if best_score > 0 else None


def is_paid_header(header: str) -> bool:
    lowered = normalize_header(header)
    paid_terms = ("плат", "договор", "средств физ", "средств юрид", "внебюдж")
    budget_terms = ("бюджет", "ассигн", "фед", "субъект", "местн")
    paid_score = sum(1 for term in paid_terms if term in lowered)
    budget_score = sum(1 for term in budget_terms if term in lowered)
    return paid_score > 0 and paid_score >= budget_score


def is_budget_header(header: str) -> bool:
    lowered = normalize_header(header)
    if is_paid_header(lowered):
        return False
    return any(term in lowered for term in ("бюджет", "ассигн", "фед", "субъект", "местн"))


def detect_per_course_groups(columns: list[str]) -> dict[str, dict[str, list[int]]]:
    groups: dict[str, dict[str, list[int]]] = {}
    for idx, column in enumerate(columns):
        lowered = normalize_header(column)
        match = re.search(r"(\d{1,2})\s*курс", lowered)
        if not match:
            continue
        course = match.group(1)
        groups.setdefault(course, {"budget": [], "paid": []})
        if is_paid_header(lowered):
            groups[course]["paid"].append(idx)
        elif is_budget_header(lowered):
            groups[course]["budget"].append(idx)
        else:
            # If per-course column has no explicit funding type label, treat as budget by default.
            groups[course]["budget"].append(idx)
    return groups


def standardize_vacancy_table(columns: list[str], rows: list[list[str]]) -> dict[str, Any] | None:
    if not columns or not rows:
        return None

    headers = [normalize_header(column) for column in columns]

    code_idx = best_index(
        columns,
        lambda h: (2 if "код" in h else 0)
        + (2 if any(term in h for term in ("специальн", "направлен", "професс", "шифр")) else 0),
    )
    direction_idx = best_index(
        columns,
        lambda h: (
            (
                3
                if ("наимен" in h and any(term in h for term in ("професс", "специальн", "направлен")))
                else 0
            )
            + (2 if any(term in h for term in ("специальн", "направлен")) else 0)
            + (
                2
                if h.strip()
                in (
                    "специальность",
                    "направление",
                    "направление подготовки",
                    "направления подготовки",
                )
                else 0
            )
            + (2 if re.fullmatch(r"наименование(?:\s+наименование)?", h.strip()) else 0)
            + (2 if h.startswith("наименование") else 0)
            + (1 if ("подготов" in h and "направлен" in h) else 0)
        )
        - (2 if "код" in h else 0)
        - (2 if any(term in h for term in ("образовательная программа", "направленност", "профил", "специализац")) else 0),
    )

    if direction_idx is None:
        direction_candidates: list[int] = []
        blocked_terms = (
            "код",
            "курс",
            "форма",
            "уровень",
            "бюдж",
            "договор",
            "плат",
            "ассигн",
            "субъект",
            "местн",
            "фед",
        )
        for idx, header in enumerate(headers):
            if "наимен" not in header:
                continue
            if any(term in header for term in blocked_terms):
                continue
            direction_candidates.append(idx)
        if direction_candidates:
            direction_idx = direction_candidates[0]
    program_idx = best_index(
        columns,
        lambda h: 3
        if any(term in h for term in ("образовательная программа", "направленност", "профил", "специализац"))
        else 0,
    )
    level_idx = best_index(columns, lambda h: 3 if ("уровень" in h and "образован" in h) else 0)
    course_idx = best_index(columns, lambda h: 3 if ("курс" in h and "курсов" not in h) else 0)
    form_idx = best_index(
        columns,
        lambda h: (
            5
            if ("форм" in h and "обуч" in h)
            else 4
            if bool(re.search(r"\bформа\b", h))
            else 3
            if h.startswith("форм")
            else 0
        ),
    )

    def column_value_ratio(idx: int, predicate: Any) -> float:
        if idx is None or idx < 0:
            return 0.0
        values = [clean_text(row[idx]) for row in rows if idx < len(row)]
        values = [value for value in values if value]
        if not values:
            return 0.0
        return sum(1 for value in values if predicate(value)) / len(values)

    if form_idx is not None:
        if column_value_ratio(form_idx, looks_like_form_value) < 0.4:
            form_idx = None

    if level_idx is not None:
        if column_value_ratio(level_idx, looks_like_level_value) < 0.3:
            level_idx = None

    form_literals = {detect_form_literal(column) for column in columns}
    form_literals.discard("")
    header_form_default = next(iter(form_literals)) if len(form_literals) == 1 else ""

    budget_indices: list[int] = []
    paid_indices: list[int] = []
    for idx, header in enumerate(headers):
        if is_paid_header(header):
            paid_indices.append(idx)
            continue
        if is_budget_header(header):
            budget_indices.append(idx)

    course_groups = detect_per_course_groups(columns)
    has_group_amount_columns = any(
        group["budget"] or group["paid"] for group in course_groups.values()
    )
    use_course_groups = bool(course_groups) and len(course_groups) >= 2 and has_group_amount_columns

    # Merge rows that contain only continuation text (often split by HTML line breaks).
    merged_rows: list[list[str]] = []
    row_list = [list(row) for row in rows]
    i = 0
    while i < len(row_list):
        row = row_list[i]
        code_value = clean_text(row[code_idx]) if code_idx is not None and code_idx < len(row) else ""
        direction_value = (
            clean_text(row[direction_idx]) if direction_idx is not None and direction_idx < len(row) else ""
        )
        program_value = (
            clean_text(row[program_idx]) if program_idx is not None and program_idx < len(row) else ""
        )

        has_other = False
        for idx, cell in enumerate(row):
            if idx in {code_idx, direction_idx, program_idx}:
                continue
            if clean_text(cell):
                has_other = True
                break

        if not code_value and (direction_value or program_value) and not has_other:
            merged = False
            if i + 1 < len(row_list):
                next_row = row_list[i + 1]
                next_code = (
                    clean_text(next_row[code_idx])
                    if code_idx is not None and code_idx < len(next_row)
                    else ""
                )
                next_direction = (
                    clean_text(next_row[direction_idx])
                    if direction_idx is not None and direction_idx < len(next_row)
                    else ""
                )
                next_program = (
                    clean_text(next_row[program_idx])
                    if program_idx is not None and program_idx < len(next_row)
                    else ""
                )
                next_has_other = False
                for idx, cell in enumerate(next_row):
                    if idx in {code_idx, direction_idx, program_idx}:
                        continue
                    if clean_text(cell):
                        next_has_other = True
                        break

                if next_code and not (next_direction or next_program) and next_has_other:
                    if direction_value and direction_idx is not None and direction_idx < len(next_row):
                        next_row[direction_idx] = clean_text(f"{direction_value} {next_row[direction_idx]}")
                    if program_value and program_idx is not None and program_idx < len(next_row):
                        next_row[program_idx] = clean_text(f"{program_value} {next_row[program_idx]}")
                    merged = True

            if not merged and merged_rows:
                target = merged_rows[-1]
                if direction_value and direction_idx is not None and direction_idx < len(target):
                    target[direction_idx] = clean_text(f"{target[direction_idx]} {direction_value}")
                if program_value and program_idx is not None and program_idx < len(target):
                    target[program_idx] = clean_text(f"{target[program_idx]} {program_value}")
            i += 1
            continue

        merged_rows.append(row)
        i += 1

    rows = merged_rows

    standardized_rows: list[list[str]] = []
    context: dict[str, str] = {
        "code": "",
        "direction": "",
        "program": "",
        "level": "",
        "course": "",
    }
    for row in rows:
        code = clean_text(row[code_idx]) if code_idx is not None and code_idx < len(row) else ""
        direction = (
            clean_text(row[direction_idx]) if direction_idx is not None and direction_idx < len(row) else ""
        )
        program = (
            clean_text(row[program_idx]) if program_idx is not None and program_idx < len(row) else ""
        )
        level = clean_text(row[level_idx]) if level_idx is not None and level_idx < len(row) else ""
        form = clean_text(row[form_idx]) if form_idx is not None and form_idx < len(row) else ""
        if not form and header_form_default:
            form = header_form_default

        if not direction and program:
            direction, program = program, ""

        split_from_code = split_code_and_direction(code)
        if split_from_code:
            split_code, split_direction = split_from_code
            if not direction or direction == code:
                code, direction = split_code, split_direction
            else:
                code = split_code

        code_valid = is_valid_code(code)
        if not code_valid:
            code = ""

        if not code and direction:
            split_from_direction = split_code_and_direction(direction)
            if split_from_direction:
                code, direction = split_from_direction

        if not code and program:
            split_from_program = split_code_and_direction(program)
            if split_from_program:
                code, program = split_from_program

        if not level and code:
            level = infer_level_from_code(code)

        course_value = (
            normalize_course(row[course_idx]) if course_idx is not None and course_idx < len(row) else ""
        )

        amount_indices: list[int] = []
        if use_course_groups:
            for group in course_groups.values():
                amount_indices.extend(group["budget"])
                amount_indices.extend(group["paid"])
        else:
            amount_indices.extend(budget_indices)
            amount_indices.extend(paid_indices)

        has_amount_values = any(
            parse_number(row[idx]) > 0 for idx in amount_indices if idx < len(row)
        )

        has_tabular_payload = bool(course_value or form or has_amount_values)
        has_key_fields = bool(code or direction or program or level)
        if not has_key_fields and has_tabular_payload and context["direction"]:
            code = context["code"]
            direction = context["direction"]
            program = context["program"]
            level = context["level"]
        elif has_tabular_payload and context["direction"]:
            if not code:
                code = context["code"]
            if not direction:
                direction = context["direction"]
            if not level:
                level = context["level"]

        if program and program == direction:
            program = ""

        def row_is_noise(check_course: str) -> bool:
            values = [code, direction, program, level, check_course, form]
            non_empty = [value for value in values if value]
            if any(is_summary_value(value) for value in non_empty):
                return True
            if len(non_empty) >= 2:
                lowered = [value.lower() for value in non_empty]
                if len(set(lowered)) == 1 and not is_valid_code(non_empty[0]):
                    return True
            numeric_like = sum(1 for value in non_empty if is_numeric_cell(value))
            if numeric_like >= 3 and not is_valid_code(code):
                return True
            if check_course and not check_course.isdigit():
                lowered_course = check_course.lower()
                if "курс" in lowered_course or is_summary_value(check_course):
                    return True
            if code and not is_valid_code(code) and len(code) > 30:
                return True
            if code and not is_valid_code(code) and is_numeric_cell(direction):
                return True
            if code and not is_valid_code(code) and not direction and not program and not level and not form:
                return True
            if (
                not code
                and direction
                and not program
                and not level
                and not form
                and not check_course
                and len(direction) > 200
            ):
                lowered_direction = direction.lower()
                menu_hits = sum(1 for word in MENU_WORDS if word in lowered_direction)
                if (
                    menu_hits >= 3
                    or "университет:" in lowered_direction
                    or "погода" in lowered_direction
                    or "liveinternet" in lowered_direction
                    or "togglehelp" in lowered_direction
                ):
                    return True
            repeated_text_values = [value for value in (direction, program, level, form, check_course) if value]
            if repeated_text_values:
                lowered_values = [value.lower() for value in repeated_text_values]
                common = max(set(lowered_values), key=lowered_values.count)
                if lowered_values.count(common) >= 3:
                    if (
                        "программы высшего образования" in common
                        or "программы среднего профессионального образования" in common
                        or "филиал " in common
                        or common.startswith("высшее образование")
                        or common.startswith("среднее профессиональное образование")
                    ):
                        return True
            return False

        if use_course_groups:
            for course in sorted(course_groups.keys(), key=lambda item: int(item)):
                group = course_groups[course]
                budget_total = sum(parse_number(row[idx]) for idx in group["budget"] if idx < len(row))
                paid_total = sum(parse_number(row[idx]) for idx in group["paid"] if idx < len(row))
                if not any((code, direction, program, level, form, budget_total, paid_total)):
                    continue
                if row_is_noise(course):
                    continue
                standardized_rows.append(
                    [
                        code,
                        direction,
                        program,
                        level,
                        course,
                        form,
                        str(budget_total),
                        str(paid_total),
                    ]
                )
            if code or direction or program or level:
                context = {
                    "code": code,
                    "direction": direction,
                    "program": program,
                    "level": level,
                    "course": context.get("course", ""),
                }
            continue

        raw_course_value = row[course_idx] if course_idx is not None and course_idx < len(row) else ""
        course_tokens = parse_course_tokens(raw_course_value)
        can_expand_by_course = False
        amount_token_map: dict[int, list[int]] = {}
        if len(course_tokens) >= 2:
            exact_match_columns = 0
            incompatible_sequence = False
            for idx in budget_indices + paid_indices:
                if idx >= len(row):
                    continue
                tokens = parse_amount_tokens(row[idx])
                amount_token_map[idx] = tokens
                if not tokens:
                    continue
                if len(tokens) == len(course_tokens):
                    exact_match_columns += 1
                    continue
                if len(tokens) > 1:
                    incompatible_sequence = True
                    break
            can_expand_by_course = exact_match_columns > 0 and not incompatible_sequence

        if can_expand_by_course:
            for pos, course_token in enumerate(course_tokens):
                budget_total = 0
                for idx in budget_indices:
                    if idx >= len(row):
                        continue
                    token_values = amount_token_map.get(idx)
                    if token_values is None:
                        token_values = parse_amount_tokens(row[idx])
                    if len(token_values) == len(course_tokens):
                        budget_total += token_values[pos]
                    elif len(token_values) == 1 and pos == 0:
                        budget_total += token_values[0]

                paid_total = 0
                for idx in paid_indices:
                    if idx >= len(row):
                        continue
                    token_values = amount_token_map.get(idx)
                    if token_values is None:
                        token_values = parse_amount_tokens(row[idx])
                    if len(token_values) == len(course_tokens):
                        paid_total += token_values[pos]
                    elif len(token_values) == 1 and pos == 0:
                        paid_total += token_values[0]

                if not any((code, direction, program, level, course_token, form, budget_total, paid_total)):
                    continue
                if row_is_noise(course_token):
                    continue
                standardized_rows.append(
                    [
                        code,
                        direction,
                        program,
                        level,
                        course_token,
                        form,
                        str(budget_total),
                        str(paid_total),
                    ]
                )

            if code or direction or program or level:
                context = {
                    "code": code,
                    "direction": direction,
                    "program": program,
                    "level": level,
                    "course": course_tokens[-1] if course_tokens else context.get("course", ""),
                }
            continue

        course = course_value
        if (
            not course
            and has_tabular_payload
            and context["course"]
            and code == context["code"]
            and direction == context["direction"]
        ):
            course = context["course"]
        budget_total = sum(parse_number(row[idx]) for idx in budget_indices if idx < len(row))
        paid_total = sum(parse_number(row[idx]) for idx in paid_indices if idx < len(row))

        if not any((code, direction, program, level, course, form, budget_total, paid_total)):
            continue
        if row_is_noise(course):
            continue

        standardized_rows.append(
            [
                code,
                direction,
                program,
                level,
                course,
                form,
                str(budget_total),
                str(paid_total),
            ]
        )
        if code or direction or program or level:
            context = {
                "code": code,
                "direction": direction,
                "program": program,
                "level": level,
                "course": course,
            }

    if not standardized_rows:
        return None

    return {"columns": STANDARD_COLUMNS, "rows": standardized_rows}


def recover_mojibake(text: str) -> str:
    if not text:
        return text

    candidates = [text]
    for source_encoding, target_encoding in (
        ("latin1", "utf-8"),
        ("latin1", "cp1251"),
        ("latin1", "koi8-r"),
        ("cp1251", "utf-8"),
        ("koi8-r", "utf-8"),
        ("cp866", "utf-8"),
    ):
        try:
            encoded = text.encode(source_encoding)
        except UnicodeError:
            continue
        try:
            converted = encoded.decode(target_encoding)
            candidates.append(converted)
        except UnicodeError:
            try:
                converted = encoded.decode(target_encoding, errors="ignore")
                if converted:
                    candidates.append(converted)
            except UnicodeError:
                continue

    best = max(candidates, key=text_quality_score)
    return best


def first_url(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    match = URL_RE.search(text)
    if not match:
        return ""
    return match.group(0)


def season_split(raw_text: str) -> dict[str, list[str]]:
    text = clean_text(raw_text)
    if not text or text.lower() in {"нет данных", "не применимо"}:
        return {"winter": [], "summer": [], "other": []}

    parts = [chunk.strip(" .") for chunk in re.split(r"[;\n]", text) if chunk.strip()]
    if not parts:
        parts = [text]

    winter_keywords = ("янв", "фев", "дек", "зим")
    summer_keywords = ("май", "июн", "июл", "авг", "лет")

    winter: list[str] = []
    summer: list[str] = []
    other: list[str] = []

    for part in parts:
        low = part.lower()
        has_winter = any(key in low for key in winter_keywords)
        has_summer = any(key in low for key in summer_keywords)
        if has_winter and not has_summer:
            winter.append(part)
        elif has_summer and not has_winter:
            summer.append(part)
        elif has_winter and has_summer:
            winter.append(part)
            summer.append(part)
        else:
            other.append(part)

    return {"winter": winter, "summer": summer, "other": other}


def table_score(columns: list[str], rows: list[list[str]]) -> float:
    sample = " ".join(columns)
    for row in rows[:15]:
        sample += " " + " ".join(row)
    sample = sample.lower()
    score = 0.0
    for key in TABLE_KEYWORDS:
        if key in sample:
            score += 6
    score += min(len(rows), 400) / 20
    score += min(len(columns), 20) / 3
    if len(columns) <= 1:
        score -= 8
    return score


def collapse_duplicate_text(text: str) -> str:
    cleaned = clean_text(text)
    if not cleaned:
        return ""
    if cleaned == "№ №":
        return "№"

    tokens = cleaned.split()
    if len(tokens) >= 2 and len(tokens) % 2 == 0:
        half = len(tokens) // 2
        if tokens[:half] == tokens[half:]:
            return " ".join(tokens[:half])

    return cleaned


def row_similarity(row: list[str], columns: list[str]) -> float:
    if not row or not columns:
        return 0.0
    max_len = max(len(row), len(columns))
    if max_len == 0:
        return 0.0
    matches = 0
    for idx in range(max_len):
        left = row[idx] if idx < len(row) else ""
        right = columns[idx] if idx < len(columns) else ""
        if left and right and left.lower() == right.lower():
            matches += 1
    return matches / max_len


def detect_data_start(rows: list[list[str]]) -> int | None:
    for idx, row in enumerate(rows[:12]):
        joined = " ".join(row)
        if CODE_RE.search(joined):
            return idx
    return None


def build_columns_from_headers(header_rows: list[list[str]]) -> list[str]:
    width = max(len(row) for row in header_rows)
    columns: list[str] = []
    for col_idx in range(width):
        parts: list[str] = []
        for row in header_rows:
            if col_idx >= len(row):
                continue
            value = collapse_duplicate_text(row[col_idx])
            if value and value not in parts:
                parts.append(value)
        name = " / ".join(parts)
        columns.append(name or f"Колонка {col_idx + 1}")
    return columns


def normalize_matrix(matrix: list[list[str]]) -> dict[str, Any] | None:
    cleaned = []
    for row in matrix:
        current = [clean_text(cell) for cell in row]
        if any(current):
            cleaned.append(current)
    if len(cleaned) < 2:
        return None

    width = max(len(row) for row in cleaned)
    normalized = [row + [""] * (width - len(row)) for row in cleaned]

    data_start = detect_data_start(normalized)
    header_blob = " ".join(" ".join(row) for row in normalized[: min(4, len(normalized))]).lower()
    if data_start is not None and data_start > 0:
        header_rows = normalized[:data_start]
        raw_columns = build_columns_from_headers(header_rows)
        data_rows = normalized[data_start:]
    else:
        raw_columns = normalized[0]
        data_rows = normalized[1:]

        if any(CODE_RE.search(cell) for cell in raw_columns):
            raw_columns = [f"Колонка {idx + 1}" for idx in range(width)]
            data_rows = normalized

    if not any(raw_columns):
        raw_columns = [f"Колонка {idx + 1}" for idx in range(width)]

    # If header looks like a scanned signature block, prefer generic headers.
    if any(hint in header_blob for hint in SIGNATURE_HINTS):
        raw_columns = [f"Колонка {idx + 1}" for idx in range(width)]

    rows = data_rows
    generic_cols = sum(
        1 for column in raw_columns if re.search(r"^колонка\s*\d+$", column.strip(), re.I)
    )
    if rows and generic_cols >= max(2, len(raw_columns) // 2):
        row0_blob = " ".join(rows[0]).lower()
        if any(hint in row0_blob for hint in HEADER_HINTS):
            raw_columns = [
                collapse_duplicate_text(cell) or f"Колонка {idx + 1}"
                for idx, cell in enumerate(rows[0])
            ]
            rows = rows[1:]

    seen: dict[str, int] = {}
    columns: list[str] = []
    for idx, column in enumerate(raw_columns, start=1):
        base = collapse_duplicate_text(column) or f"Колонка {idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        if count > 1:
            base = f"{base} ({count})"
        columns.append(base)

    if rows and row_similarity(rows[0], columns) > 0.7:
        rows = rows[1:]

    if rows:
        keep = [idx for idx in range(len(columns)) if any(row[idx] for row in rows)]
        if keep and len(keep) < len(columns):
            columns = [columns[idx] for idx in keep]
            rows = [[row[idx] for idx in keep] for row in rows]

    columns = apply_column_inference(columns, rows)

    return {"columns": columns, "rows": rows}


def infer_column_name(values: list[str]) -> str | None:
    sample = [value for value in values if value]
    if not sample:
        return None

    lowered = [value.lower() for value in sample]
    total = len(lowered)

    def ratio(predicate: Any) -> float:
        return sum(1 for value in lowered if predicate(value)) / total

    if ratio(lambda value: "курс" in value) >= 0.45:
        return "Курс"
    if ratio(
        lambda value: any(
            hint in value
            for hint in (
                "очная",
                "заочная",
                "очно-заочная",
                "очно заочная",
                "дистанц",
                "вечерн",
            )
        )
    ) >= 0.45:
        return "Форма"
    if ratio(lambda value: bool(CODE_RE.search(value))) >= 0.4:
        return "Код"
    if ratio(
        lambda value: any(
            hint in value
            for hint in (
                "бакалав",
                "магистр",
                "специал",
                "аспиран",
                "ординат",
                "среднее профессион",
                "спо",
            )
        )
    ) >= 0.35:
        return "Уровень"

    avg_len = sum(len(value) for value in sample) / total
    if avg_len >= 18 and ratio(lambda value: bool(CYRILLIC_RE.search(value))) >= 0.6:
        return "Направление"

    upper_hits = 0
    for value in sample:
        letters = re.sub(r"[^A-Za-zА-Яа-яЁё]", "", value)
        if letters and letters == letters.upper():
            upper_hits += 1
    if upper_hits / total >= 0.6 and avg_len <= 8:
        return "Подразделение"

    return None


def apply_column_inference(columns: list[str], rows: list[list[str]]) -> list[str]:
    if not columns or not rows:
        return columns

    if len(columns) == 24 and rows:
        sample_row = rows[0]
        if len(sample_row) == 24 and looks_like_form_value(sample_row[1]) and CODE_RE.search(sample_row[2]):
            base = ["Подразделение", "Форма", "Код", "Направление"]
            course_cols = []
            for course in range(1, 6):
                for label in (
                    "Фед. бюджет",
                    "Бюджет субъекта РФ",
                    "Местный бюджет",
                    "Платное обучение",
                ):
                    course_cols.append(f"{course} курс — {label}")
            return base + course_cols

    updated = columns[:]
    for idx, column in enumerate(columns):
        if not re.search(r"^колонка\s*\d+$", column.strip(), re.I):
            continue
        inferred = infer_column_name([row[idx] for row in rows if idx < len(row)])
        if inferred:
            updated[idx] = inferred

    return updated


def looks_like_form_value(value: str) -> bool:
    lowered = value.lower()
    return any(
        hint in lowered
        for hint in (
            "очная",
            "заочная",
            "очно-заочная",
            "очно заочная",
            "дистанц",
            "вечерн",
        )
    )


def extract_cell_text(cell: Any) -> str:
    hidden_texts = [
        clean_text(hidden.get_text(" ", strip=True))
        for hidden in cell.select(HIDDEN_CELL_SELECTORS)
    ]
    for hidden in cell.select(HIDDEN_CELL_SELECTORS):
        hidden.extract()

    text = clean_text(cell.get_text(" ", strip=True))
    if not text:
        image = cell.find("img")
        if image:
            text = clean_text(image.get("alt") or image.get("title") or "")
    if not text:
        text = clean_text(cell.get("aria-label") or "")
    if not text and hidden_texts:
        text = " ".join(value for value in hidden_texts if value)
    return text


def is_hidden_cell(cell: Any) -> bool:
    classes = set(cell.get("class", []) or [])
    if classes & HIDDEN_CELL_CLASSES:
        return True
    style = (cell.get("style") or "").lower()
    if "display:none" in style or "visibility:hidden" in style:
        return True
    return False


def parse_html_table(table_tag: Any) -> dict[str, Any] | None:
    matrix: list[list[str]] = []
    spans: dict[int, dict[str, Any]] = {}

    for tr in table_tag.find_all("tr"):
        if tr.find_parent("table") is not table_tag:
            continue
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue

        row: list[str] = []
        col_idx = 0

        def fill_spans() -> None:
            nonlocal col_idx
            while col_idx in spans:
                span = spans[col_idx]
                row.append(span["value"])
                span["rows_left"] -= 1
                if span["rows_left"] <= 0:
                    spans.pop(col_idx, None)
                col_idx += 1

        for cell in cells:
            if is_hidden_cell(cell):
                continue
            fill_spans()
            value = extract_cell_text(cell)
            colspan = int(cell.get("colspan", 1) or 1)
            rowspan = int(cell.get("rowspan", 1) or 1)
            for _ in range(colspan):
                row.append(value)
                if rowspan > 1:
                    spans[col_idx] = {"value": value, "rows_left": rowspan - 1}
                col_idx += 1

        fill_spans()

        if any(row):
            matrix.append(row)

    return normalize_matrix(matrix)


def table_metrics(columns: list[str], rows: list[list[str]]) -> dict[str, float]:
    sample = " ".join(" ".join(row) for row in rows[:12]).lower()
    digit_ratio = sum(ch.isdigit() for ch in sample) / (len(sample) or 1)
    menu_hits = sum(sample.count(word) for word in MENU_WORDS)
    generic_cols = sum(
        1 for column in columns if re.search(r"^колонка\s*\d+$", column.strip(), re.I)
    )
    unnamed_cols = sum(1 for column in columns if "unnamed" in column.lower())
    max_col_len = max((len(column) for column in columns), default=0)
    header_hits = sum(
        1 for column in columns if any(key in column.lower() for key in ("код", "наименован", "курс", "форма", "уровень", "бюдж"))
    )
    return {
        "digit_ratio": digit_ratio,
        "menu_hits": float(menu_hits),
        "generic_cols": float(generic_cols),
        "unnamed_cols": float(unnamed_cols),
        "max_col_len": float(max_col_len),
        "header_hits": float(header_hits),
    }


def score_table(columns: list[str], rows: list[list[str]]) -> float:
    base = table_score(columns, rows)
    metrics = table_metrics(columns, rows)
    score = base
    score += metrics["digit_ratio"] * 40
    score += metrics["header_hits"] * 2.5
    score -= metrics["menu_hits"] * 3
    score -= metrics["generic_cols"] * 5
    score -= metrics["unnamed_cols"] * 8
    if columns:
        generic_ratio = metrics["generic_cols"] / len(columns)
        score -= generic_ratio * 40
        if generic_ratio > 0.4:
            score -= 15
    if len(columns) > 60:
        score -= 25
    if metrics["max_col_len"] > 280:
        score -= 12
    return score


def is_layout_table(columns: list[str], rows: list[list[str]]) -> bool:
    if len(columns) > 80:
        return True
    if len(rows) < 2:
        return True
    metrics = table_metrics(columns, rows)
    if metrics["digit_ratio"] < 0.01 and metrics["menu_hits"] >= 2:
        return True
    if metrics["generic_cols"] >= max(2, len(columns) // 2) and metrics["digit_ratio"] < 0.02:
        return True
    if metrics["max_col_len"] > 500 and len(columns) <= 2:
        return True
    return False


def collect_download_links(soup: BeautifulSoup, base_url: str) -> list[str]:
    candidates: dict[str, float] = {}
    for anchor in soup.find_all("a", href=True):
        url = urljoin(base_url, anchor["href"])
        parsed = urlparse(url)
        suffix = Path(parsed.path).suffix.lower()
        text = (clean_text(anchor.get_text(" ", strip=True)) + " " + url).lower()
        if suffix in FILE_EXTENSIONS or any(key in text for key in DOWNLOAD_KEYWORDS):
            score = 0.0
            if suffix in FILE_EXTENSIONS:
                score += 2
            for key in DOWNLOAD_KEYWORDS:
                if key in text:
                    score += 4
            current = candidates.get(url, -1)
            if score > current:
                candidates[url] = score
    ordered = sorted(candidates.items(), key=lambda item: item[1], reverse=True)
    return [url for url, _ in ordered[:10]]


def extract_html_best_table(html: str, base_url: str) -> tuple[dict[str, Any] | None, list[str]]:
    if not BS4_AVAILABLE:
        return None, []
    soup = BeautifulSoup(html, "lxml")
    tables: list[dict[str, Any]] = []
    for table_tag in soup.find_all("table")[:40]:
        parsed = parse_html_table(table_tag)
        if not parsed:
            continue
        if is_layout_table(parsed["columns"], parsed["rows"]):
            continue
        parsed["score"] = score_table(parsed["columns"], parsed["rows"])
        tables.append(parsed)
    best = max(tables, key=lambda table: table["score"]) if tables else None
    links = collect_download_links(soup, base_url)
    return best, links


def extract_pandas_html_table(content: bytes) -> dict[str, Any] | None:
    if pd is None:
        return None
    candidates: list[dict[str, Any]] = []
    try:
        dataframes = pd.read_html(io.BytesIO(content))
    except (ValueError, ImportError):
        return None
    except Exception:  # noqa: BLE001
        return None

    for dataframe in dataframes[:20]:
        try:
            df = dataframe.copy()
            df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
            if df.empty:
                continue

            columns = []
            for idx, column in enumerate(df.columns.tolist(), start=1):
                if isinstance(column, tuple):
                    joined = " ".join(clean_text(part) for part in column if clean_text(part))
                    columns.append(joined or f"Колонка {idx}")
                else:
                    columns.append(clean_text(column) or f"Колонка {idx}")

            rows = [[clean_text(cell) for cell in row] for row in df.fillna("").values.tolist()]
            columns = apply_column_inference(columns, rows)
            if is_layout_table(columns, rows):
                continue
            candidates.append({"columns": columns, "rows": rows, "score": score_table(columns, rows)})
        except Exception:  # noqa: BLE001
            continue

    if not candidates:
        return None
    return max(candidates, key=lambda item: item["score"])


def decode_html(content: bytes, declared_encoding: str | None) -> str:
    preferred_encodings = []
    if declared_encoding:
        preferred_encodings.append(declared_encoding)
    if UnicodeDammit is not None:
        guessed = UnicodeDammit(content, preferred_encodings).unicode_markup
        if guessed:
            return guessed
    elif declared_encoding:
        try:
            return content.decode(declared_encoding, errors="replace")
        except LookupError:
            pass
    return content.decode("utf-8", errors="replace")


def parse_pdf_table(content: bytes) -> dict[str, Any] | None:
    if pdfplumber is None:
        return None
    pdf_candidates: list[dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages[:30]:
            for raw_table in page.extract_tables() or []:
                matrix = raw_table or []
                parsed = normalize_matrix(matrix)
                if not parsed:
                    continue
                if is_layout_table(parsed["columns"], parsed["rows"]):
                    continue
                parsed["score"] = score_table(parsed["columns"], parsed["rows"])
                parsed["_order"] = (page.page_number * 1000) + len(pdf_candidates)
                pdf_candidates.append(parsed)

    merged_pdf = merge_similar_tables(pdf_candidates)
    if merged_pdf:
        return merged_pdf

    try:
        import camelot  # type: ignore
        import tempfile
        import os

        fd, path = tempfile.mkstemp(suffix=".pdf")
        os.write(fd, content)
        os.close(fd)
        try:
            camelot_candidates: list[dict[str, Any]] = []
            for flavor in ("lattice", "stream"):
                tables = camelot.read_pdf(path, pages="1-3", flavor=flavor)
                for table_idx, table in enumerate(tables):
                    matrix = table.df.values.tolist()
                    parsed = normalize_matrix(matrix)
                    if not parsed:
                        continue
                    if is_layout_table(parsed["columns"], parsed["rows"]):
                        continue
                    parsed["score"] = score_table(parsed["columns"], parsed["rows"])
                    parsed["_order"] = (1_000_000 if flavor == "lattice" else 2_000_000) + table_idx
                    camelot_candidates.append(parsed)
            merged_camelot = merge_similar_tables(camelot_candidates)
            if merged_camelot:
                return merged_camelot
        finally:
            os.remove(path)
    except Exception:
        pass

    return None


def parse_excel_table(content: bytes) -> dict[str, Any] | None:
    if pd is None:
        return None
    stream = io.BytesIO(content)
    xls = pd.ExcelFile(stream)
    candidates: list[dict[str, Any]] = []

    for sheet in xls.sheet_names[:8]:
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
        df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if df.empty:
            continue
        columns = [
            clean_text(column) if clean_text(column) else f"Колонка {idx + 1}"
            for idx, column in enumerate(df.columns.tolist())
        ]
        rows = [[clean_text(cell) for cell in row] for row in df.fillna("").values.tolist()]
        columns = apply_column_inference(columns, rows)
        score = table_score(columns, rows)
        candidates.append({"columns": columns, "rows": rows, "score": score})

    if not candidates:
        return None
    return max(candidates, key=lambda item: item["score"])


def parse_csv_table(content: bytes) -> dict[str, Any] | None:
    if pd is None:
        return None
    stream = io.BytesIO(content)
    df = pd.read_csv(stream, dtype=str, sep=None, engine="python")
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty:
        return None
    columns = [
        clean_text(column) if clean_text(column) else f"Колонка {idx + 1}"
        for idx, column in enumerate(df.columns.tolist())
    ]
    rows = [[clean_text(cell) for cell in row] for row in df.fillna("").values.tolist()]
    columns = apply_column_inference(columns, rows)
    return {"columns": columns, "rows": rows, "score": table_score(columns, rows)}


def response_is_html(url: str, response: Response) -> bool:
    content_type = response.headers.get("Content-Type", "").lower()
    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in FILE_EXTENSIONS:
        return False
    return "html" in content_type or "text/" in content_type or not content_type


def parse_binary_table(url: str, response: Response) -> dict[str, Any] | None:
    suffix = Path(urlparse(url).path).suffix.lower()
    content_type = response.headers.get("Content-Type", "").lower()

    if suffix == ".pdf" or "pdf" in content_type:
        return parse_pdf_table(response.content)
    if suffix in {".xls", ".xlsx"} or "spreadsheet" in content_type or "excel" in content_type:
        return parse_excel_table(response.content)
    if suffix == ".csv" or "csv" in content_type:
        return parse_csv_table(response.content)
    return None


def table_header_similarity(left: list[str], right: list[str]) -> float:
    if not left or not right:
        return 0.0
    left_norm = [normalize_header(item) for item in left]
    right_norm = [normalize_header(item) for item in right]

    positional = 0
    width = max(len(left_norm), len(right_norm))
    for idx in range(min(len(left_norm), len(right_norm))):
        if left_norm[idx] and left_norm[idx] == right_norm[idx]:
            positional += 1
    positional_score = positional / width if width else 0.0

    left_set = {item for item in left_norm if item}
    right_set = {item for item in right_norm if item}
    if not left_set or not right_set:
        token_score = 0.0
    else:
        token_score = len(left_set & right_set) / len(left_set | right_set)
    return max(positional_score, token_score)


def header_kind(header: str) -> str:
    lowered = normalize_header(header)
    if not lowered:
        return "unknown"
    if "п/п" in lowered or lowered in {"№", "номер"}:
        return "ordinal"
    if is_paid_header(lowered):
        return "paid"
    if "субъект" in lowered:
        return "budget_subject"
    if "местн" in lowered:
        return "budget_local"
    if is_budget_header(lowered):
        if "фед" in lowered or "ассигн" in lowered:
            return "budget_federal"
        return "budget_other"
    if "курс" in lowered and "курсов" not in lowered:
        return "course"
    if ("форм" in lowered and "обуч" in lowered) or re.search(r"\bформа\b", lowered):
        return "form"
    if "уровень" in lowered and "образован" in lowered:
        return "level"
    if any(term in lowered for term in ("образовательная программа", "направленност", "профил", "специализац")):
        return "program"
    if "код" in lowered:
        return "code"
    if "наимен" in lowered and any(term in lowered for term in ("специальн", "направлен", "професс")):
        return "direction"
    if lowered.startswith("наименование"):
        return "direction"
    return "unknown"


def align_row_to_columns(row: list[str], source_columns: list[str], target_columns: list[str]) -> list[str]:
    aligned = [""] * len(target_columns)
    source_kinds = [header_kind(header) for header in source_columns]
    target_kinds = [header_kind(header) for header in target_columns]

    target_positions: dict[str, list[int]] = {}
    for idx, kind in enumerate(target_kinds):
        target_positions.setdefault(kind, []).append(idx)

    for source_idx, value in enumerate(row):
        if source_idx >= len(source_columns):
            continue
        normalized_value = clean_text(value)
        if not normalized_value:
            continue

        source_kind = source_kinds[source_idx]
        target_idx: int | None = None

        if source_kind in target_positions and target_positions[source_kind]:
            source_header = normalize_header(source_columns[source_idx])
            candidates = target_positions[source_kind]
            if len(candidates) == 1:
                target_idx = candidates[0]
            else:
                best_score = -1
                for candidate_idx in candidates:
                    candidate_header = normalize_header(target_columns[candidate_idx])
                    score = 0
                    for token in ("фед", "субъект", "местн", "плат", "договор", "курс", "форма", "уровень"):
                        if token in source_header and token in candidate_header:
                            score += 2
                    if source_header == candidate_header:
                        score += 3
                    if score > best_score:
                        best_score = score
                        target_idx = candidate_idx

        if target_idx is None:
            source_header = normalize_header(source_columns[source_idx])
            for idx, target_header in enumerate(target_columns):
                if source_header and source_header == normalize_header(target_header):
                    target_idx = idx
                    break

        if target_idx is None and source_idx < len(aligned):
            target_idx = source_idx

        if target_idx is None or target_idx >= len(aligned):
            continue

        if aligned[target_idx]:
            if normalized_value not in aligned[target_idx]:
                aligned[target_idx] = clean_text(f"{aligned[target_idx]} {normalized_value}")
        else:
            aligned[target_idx] = normalized_value

    return aligned


def merge_similar_tables(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None

    anchor = max(candidates, key=lambda item: item.get("score", 0.0))
    anchor_columns = anchor.get("columns") or []
    if not anchor_columns:
        return None

    mergeable: list[dict[str, Any]] = []
    for candidate in candidates:
        columns = candidate.get("columns") or []
        if table_header_similarity(anchor_columns, columns) >= 0.55:
            mergeable.append(candidate)

    if not mergeable:
        mergeable = [anchor]

    mergeable.sort(key=lambda item: int(item.get("_order", 10_000_000)))
    merged_rows: list[list[str]] = []
    seen_rows: set[tuple[str, ...]] = set()
    for candidate in mergeable:
        candidate_columns = candidate.get("columns") or []
        for row in candidate.get("rows") or []:
            normalized = [clean_text(cell) for cell in row]
            if not any(normalized):
                continue
            if candidate_columns and len(candidate_columns) != len(anchor_columns):
                normalized = align_row_to_columns(normalized, candidate_columns, anchor_columns)
            if row_similarity(normalized, anchor_columns) > 0.8:
                continue
            key = tuple(normalized)
            if key in seen_rows:
                continue
            seen_rows.add(key)
            merged_rows.append(normalized)

    if not merged_rows:
        return None

    return {
        "columns": anchor_columns,
        "rows": merged_rows,
        "score": max(item.get("score", 0.0) for item in mergeable),
    }


def fetch(url: str) -> Response:
    last_error: Exception | None = None
    try:
        return SESSION.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
    except SSLError:
        try:
            return SESSION.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True, verify=False)
        except RequestException as exc:
            last_error = exc
    except RequestException as exc:
        last_error = exc

    if url.startswith("https://"):
        fallback_url = "http://" + url[len("https://") :]
        try:
            return SESSION.get(fallback_url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        except RequestException as exc:
            last_error = exc

    if last_error:
        raise last_error
    raise RequestException("Не удалось выполнить HTTP-запрос.")


def trim_rows(rows: list[list[str]]) -> tuple[list[list[str]], bool]:
    if len(rows) <= MAX_ROWS_PER_TABLE:
        return rows, False
    return rows[:MAX_ROWS_PER_TABLE], True


def serialize_table(parsed: dict[str, Any], source_url: str) -> dict[str, Any]:
    rows, truncated = trim_rows(parsed["rows"])
    standardized = standardize_vacancy_table(parsed["columns"], rows)
    return {
        "status": "ok",
        "message": "",
        "source_url": source_url,
        "columns": parsed["columns"],
        "rows": rows,
        "standardized_columns": standardized["columns"] if standardized else [],
        "standardized_rows": standardized["rows"] if standardized else [],
        "row_count": len(parsed["rows"]),
        "truncated": truncated,
    }


def parse_response_for_table(response: Response) -> tuple[dict[str, Any] | None, list[str], bool]:
    if response_is_html(response.url, response):
        html = decode_html(response.content, response.encoding)
        parsed_bs4, links = extract_html_best_table(html, response.url)
        parsed_pandas = extract_pandas_html_table(response.content)

        if parsed_bs4 and parsed_pandas:
            parsed_table = parsed_bs4
            if parsed_pandas.get("score", 0) > parsed_bs4.get("score", 0):
                parsed_table = parsed_pandas
        else:
            parsed_table = parsed_bs4 or parsed_pandas

        return parsed_table, links, True
    parsed_table = parse_binary_table(response.url, response)
    return parsed_table, [], False


def manual_fallback_payload(
    fallback: dict[str, Any], official_url: str, tried_urls: list[str], timestamp: str
) -> dict[str, Any]:
    rows = [[clean_text(cell) for cell in row] for row in fallback["rows"]]
    standardized = standardize_vacancy_table(fallback["columns"], rows)
    payload = {
        "status": "ok",
        "message": clean_text(fallback.get("message")),
        "source_url": clean_text(fallback.get("source_url")) or official_url,
        "columns": [clean_text(column) for column in fallback["columns"]],
        "rows": rows,
        "standardized_columns": standardized["columns"] if standardized else [],
        "standardized_rows": standardized["rows"] if standardized else [],
        "row_count": len(rows),
        "truncated": False,
        "tried_urls": tried_urls,
        "fetched_at": timestamp,
        "is_manual_fallback": True,
    }
    return payload


def extract_vacancies(official_url: str) -> dict[str, Any]:
    timestamp = datetime.now(timezone.utc).isoformat()
    if not official_url:
        return {
            "status": "error",
            "message": "Не указана ссылка на официальный источник.",
            "source_url": "",
            "columns": [],
            "rows": [],
            "row_count": 0,
            "truncated": False,
            "tried_urls": [],
            "fetched_at": timestamp,
        }

    tried_urls: list[str] = []
    seen_urls: set[str] = set()

    def add_url(url: str) -> None:
        if url and url not in seen_urls:
            seen_urls.add(url)
            tried_urls.append(url)

    add_url(official_url)

    try:
        response = fetch(official_url)
    except RequestException as exc:
        return {
            "status": "error",
            "message": f"Ошибка запроса: {exc}",
            "source_url": official_url,
            "columns": [],
            "rows": [],
            "row_count": 0,
            "truncated": False,
            "tried_urls": tried_urls,
            "fetched_at": timestamp,
        }

    add_url(response.url)
    parsed_table, links, is_html = parse_response_for_table(response)
    if parsed_table:
        payload = serialize_table(parsed_table, response.url)
        payload["tried_urls"] = tried_urls
        payload["fetched_at"] = timestamp
        return payload

    queue: list[str] = list(links)
    while queue and len(tried_urls) < 25:
        link = queue.pop(0)
        if link in seen_urls:
            continue
        add_url(link)
        try:
            linked_response = fetch(link)
        except RequestException:
            continue

        add_url(linked_response.url)
        parsed_linked, nested_links, linked_is_html = parse_response_for_table(linked_response)
        if parsed_linked:
            payload = serialize_table(parsed_linked, linked_response.url)
            payload["tried_urls"] = tried_urls
            payload["fetched_at"] = timestamp
            payload["message"] = (
                "Таблица извлечена из дополнительной ссылки, найденной на странице вуза."
            )
            return payload

        if linked_is_html:
            for nested in nested_links:
                if nested not in seen_urls and len(queue) < 50:
                    queue.append(nested)

    if is_html:
        return {
            "status": "partial",
            "message": "На странице и связанных материалах не найдена табличная разметка с вакантными местами.",
            "source_url": response.url,
            "columns": [],
            "rows": [],
            "row_count": 0,
            "truncated": False,
            "tried_urls": tried_urls,
            "fetched_at": timestamp,
        }

    return {
        "status": "partial",
        "message": "Источник открыт, но формат таблицы не удалось автоматически распознать.",
        "source_url": response.url,
        "columns": [],
        "rows": [],
        "row_count": 0,
        "truncated": False,
        "tried_urls": tried_urls,
        "fetched_at": timestamp,
    }


def read_universities(source_path: Path) -> list[dict[str, Any]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read the Excel source file. Install it with: python3 -m pip install openpyxl")
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    universities: list[dict[str, Any]] = []

    def cell_link(worksheet: Any, row: int, column: int) -> str:
        cell = worksheet.cell(row, column)
        if cell.hyperlink and cell.hyperlink.target:
            return clean_text(cell.hyperlink.target)
        return first_url(cell.value)

    def parse_dormitory_flag(raw: str) -> bool | None:
        lowered = clean_text(raw).lower()
        if not lowered:
            return None
        if lowered in {"+", "++", "есть", "да"}:
            return True
        if lowered in {"-", "нет", "no"}:
            return False
        return None

    if "Даты" in workbook.sheetnames:
        worksheet = workbook["Даты"]
        for row in range(2, worksheet.max_row + 1):
            university_id = f"u{row:03d}"
            university_name = clean_text(worksheet.cell(row, 1).value)
            if not university_name:
                continue

            marker = clean_text(worksheet.cell(row, 14).value)
            if marker in EXCLUDED_MARKERS:
                continue

            official_link = cell_link(worksheet, row, 10)
            if not official_link:
                official_link = first_url(worksheet.cell(row, 7).value)
            official_link = SOURCE_OVERRIDES.get(university_id, official_link)

            start_raw = clean_text(worksheet.cell(row, 4).value)
            universities.append(
                {
                    "id": university_id,
                    "excel_row": row,
                    "university": university_name,
                    "marker": marker,
                    "years": clean_text(worksheet.cell(row, 2).value),
                    "vacancy_publication_dates": clean_text(worksheet.cell(row, 3).value),
                    "application_start_dates_raw": start_raw,
                    "application_start_dates_by_season": season_split(start_raw),
                    "transfer_waves_description": clean_text(worksheet.cell(row, 5).value),
                    "transfer_conditions": clean_text(worksheet.cell(row, 6).value),
                    "source_reference": clean_text(worksheet.cell(row, 7).value),
                    "contacts": clean_text(worksheet.cell(row, 8).value),
                    "notes": clean_text(worksheet.cell(row, 9).value),
                    "official_vacancies_link": official_link,
                    "dormitory": {
                        "has_dormitory": parse_dormitory_flag(marker),
                        "details": "",
                        "placeholder": "Секция подготовлена, данные будут добавлены позже.",
                    },
                    "vacancies": {},
                    "allow_manual_fallback": True,
                }
            )
        return universities

    worksheet = workbook[workbook.sheetnames[0]]
    headers = {
        idx: normalize_header(worksheet.cell(1, idx).value)
        for idx in range(1, worksheet.max_column + 1)
    }

    def find_column(*keywords: str) -> int | None:
        for idx, header in headers.items():
            if all(keyword in header for keyword in keywords):
                return idx
        return None

    number_col = find_column("№") or find_column("n")
    university_col = find_column("краткое", "название", "вуз") or find_column("название", "вуз")
    dates_col = find_column("сроки", "подачи")
    link_col = find_column("ссылка", "вакант")
    transfer_notes_col = find_column("примечания", "дате")
    dormitory_col = find_column("общежитие")
    dormitory_comment_col = find_column("комментарий", "общежити")
    procedure_col = find_column("процедура", "перевод")
    docs_col = find_column("необходимые", "документ")

    if university_col is None:
        raise RuntimeError(
            f"Не удалось распознать структуру Excel: не найдена колонка с названием вуза в листе '{worksheet.title}'."
        )

    sequence = 0
    for row in range(2, worksheet.max_row + 1):
        university_name = clean_text(worksheet.cell(row, university_col).value)
        if not university_name:
            continue
        sequence += 1
        university_id = f"u{sequence:03d}"

        marker = clean_text(worksheet.cell(row, dormitory_col).value) if dormitory_col else ""
        official_link = cell_link(worksheet, row, link_col) if link_col else ""
        official_link = SOURCE_OVERRIDES.get(university_id, official_link)

        start_raw = clean_text(worksheet.cell(row, dates_col).value) if dates_col else ""
        transfer_notes = clean_text(worksheet.cell(row, transfer_notes_col).value) if transfer_notes_col else ""
        procedure = clean_text(worksheet.cell(row, procedure_col).value) if procedure_col else ""
        docs = clean_text(worksheet.cell(row, docs_col).value) if docs_col else ""
        dorm_comment = (
            clean_text(worksheet.cell(row, dormitory_comment_col).value) if dormitory_comment_col else ""
        )

        universities.append(
            {
                "id": university_id,
                "excel_row": row,
                "university": university_name,
                "marker": marker,
                "years": "",
                "vacancy_publication_dates": "",
                "application_start_dates_raw": start_raw,
                "application_start_dates_by_season": season_split(start_raw),
                "transfer_waves_description": transfer_notes or start_raw,
                "transfer_conditions": procedure,
                "source_reference": docs,
                "contacts": "",
                "notes": docs,
                "official_vacancies_link": official_link,
                "dormitory": {
                    "has_dormitory": parse_dormitory_flag(marker),
                    "details": dorm_comment,
                    "placeholder": "Секция подготовлена, данные будут добавлены позже.",
                },
                "vacancies": {},
                "allow_manual_fallback": True,
                "list_number": clean_text(worksheet.cell(row, number_col).value) if number_col else "",
            }
        )

    return universities


def enrich_with_vacancies(universities: list[dict[str, Any]], workers: int) -> None:
    total = len(universities)
    done = 0
    with futures.ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_index = {
            executor.submit(extract_vacancies, university["official_vacancies_link"]): index
            for index, university in enumerate(universities)
        }
        for completed in futures.as_completed(future_to_index):
            index = future_to_index[completed]
            try:
                result = completed.result()
            except Exception as exc:  # noqa: BLE001
                result = {
                    "status": "error",
                    "message": f"Внутренняя ошибка парсинга: {exc}",
                    "source_url": universities[index]["official_vacancies_link"],
                    "columns": [],
                    "rows": [],
                    "row_count": 0,
                    "truncated": False,
                    "tried_urls": [universities[index]["official_vacancies_link"]],
                    "fetched_at": datetime.now(timezone.utc).isoformat(),
                }

            fallback = MANUAL_VACANCY_FALLBACKS.get(universities[index]["id"])
            standardized_rows = result.get("standardized_rows")
            has_standardized_rows = isinstance(standardized_rows, list) and len(standardized_rows) > 0
            needs_fallback = result.get("status") != "ok" or not has_standardized_rows
            if needs_fallback and fallback and universities[index].get("allow_manual_fallback", True):
                result = manual_fallback_payload(
                    fallback,
                    universities[index]["official_vacancies_link"],
                    result.get("tried_urls", [universities[index]["official_vacancies_link"]]),
                    result.get("fetched_at", datetime.now(timezone.utc).isoformat()),
                )

            universities[index]["vacancies"] = result
            done += 1
            name = universities[index]["university"][:46]
            print(f"[{done:>2}/{total}] {name}", file=sys.stderr)


def build_payload(universities: list[dict[str, Any]], source_path: Path) -> dict[str, Any]:
    ok_count = sum(1 for university in universities if university["vacancies"]["status"] == "ok")
    partial_count = sum(
        1 for university in universities if university["vacancies"]["status"] == "partial"
    )
    error_count = sum(
        1 for university in universities if university["vacancies"]["status"] == "error"
    )

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": str(source_path),
        "total_universities": len(universities),
        "stats": {
            "vacancies_ok": ok_count,
            "vacancies_partial": partial_count,
            "vacancies_error": error_count,
        },
        "universities": universities,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build JSON dataset for transfer vacancies site.")
    parser.add_argument("--source", required=True, help="Path to dataset_spb_transfer_filled.xlsx")
    parser.add_argument("--out", required=True, help="Output JSON file path")
    parser.add_argument("--workers", type=int, default=8, help="Parallel workers for URL parsing")
    args = parser.parse_args()

    source_path = Path(args.source).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    universities = read_universities(source_path)
    enrich_with_vacancies(universities, workers=max(1, args.workers))
    payload = build_payload(universities, source_path)

    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"Saved: {out_path} | universities={payload['total_universities']} "
        f"| ok={payload['stats']['vacancies_ok']} "
        f"| partial={payload['stats']['vacancies_partial']} "
        f"| error={payload['stats']['vacancies_error']}"
    )


if __name__ == "__main__":
    main()
